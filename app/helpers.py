from openpyxl import load_workbook

# defining global vars

# base_filepath = '/home/jberusch/GCYC/app/resources/' # use if CSIL
base_filepath = '/home/jberusch/Desktop/GCYC/app/resources/' # use if laptop

wb1 = load_workbook(base_filepath + 'eoy_dashboard.xlsx')
wb2 = load_workbook(base_filepath + 'progress_to_date_edit.xlsx')
prog = wb2['Sheet1']

# get all column headers
def get_col_headers():
	column_headers = []
	for cell in prog[1]:
		column_headers.append(cell.value)

	return column_headers

# find row corresponding to student id input 
def get_row(id):
	for row in prog.rows:
		if (row[0].value == id):
			return row

# print a row with labels
def print_row(headers,row):
	for i in range(len(headers)):
		print(headers[i])
		print(row[i].value)

def get_filter_data(form_data, student_row, col_headers):
	results = {}

	# for every filter, get the value if it's true
	if form_data['gender']:
		results['Gender'] = student_row[col_headers.index('Gender')].value
	if form_data['ms']:
		results['Middle School'] = student_row[col_headers.index('Middle School')].value
	if form_data['school']:
		results['School'] = student_row[col_headers.index('School')].value
	if form_data['advisor']:
		results['Advisor'] = student_row[col_headers.index('Advisor')].value

	return results

def get_gpa_data(student_row, col_headers):
	res = {} # dictionary to hold all results

	# get all columns between 'Name' and 'GPA'
	index = col_headers.index('Name')
	i = 1
	current_header = 'Name'
	while current_header != 'GPA':
		current_header = col_headers[index+i]
		res[current_header] = student_row[index+i].value
		i += 1

	return res

def get_individual_student_data(form_data):
	print("========== starting get_data ==============")

	# NOTE: probably want to expand so users can input multiple IDs
	id = int(form_data['student_id'])

	col_headers = get_col_headers()
	print(col_headers) # check

	student_row = get_row(id)
	# check that a match was found
	if not student_row:
		return "Sorry, we couldn't find data on that student ID."

	print('===========================================')
	print_row(col_headers,student_row)

	# get data asked for by filters
	results = get_filter_data(form_data, student_row, col_headers)

	# always get GPA (users say they look @ that usually in chief)
	# get GPA last --> translates to "1st" in dictionary
	results['GPA Dict'] = get_gpa_data(student_row, col_headers)

	print("=========== ending get_data ===============")
	return results