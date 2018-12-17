import os
from app import app
from openpyxl import load_workbook

base_filepath = os.path.join(app.root_path,'resources/')
wb1 = load_workbook(base_filepath + 'eoy_dashboard.xlsx')
wb2 = load_workbook(base_filepath + 'progress_to_date.xlsx')
prog = wb2['Progress to Date']

# get all column headers
def get_col_headers():
	column_headers = []
	for cell in prog[1]:
		column_headers.append(cell.value)

	return column_headers

# find row corresponding to student id input 
def get_row(student_id,col_headers):
	for row in prog.rows:
		if (row[col_headers.index('Student ID')].value == student_id):
			return row

# print a row with labels
def print_row(headers,row):
	for i in range(len(headers)):
		print(headers[i])
		print(row[i].value)

# get all rows for which val is in col
def get_rows_in_group(col_headers, col_name, val):
	res = []
	col_index = col_headers.index(col_name)

	# loop through all rows
	for row in prog.rows:
		# check value in column
		if (row[col_index].value == val):
			res.append(row)

	return res